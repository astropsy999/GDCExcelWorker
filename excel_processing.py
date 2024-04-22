import os
import logging
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.cell import MergedCell
import re
from datetime import datetime
import shutil
from process_excel_rgf import process_excel_file_rgf
from process_excel_oen import process_excel_file_oen


# Функция для получения списка всех файлов Excel в папке
def get_excel_files(directory):
    excel_files = [file for file in os.listdir(directory) if file.lower().endswith(".xlsx")]
    logging.info(f'Найдены файлы Excel: {excel_files}')
    return excel_files

# Объединенная функция для обработки всех файлов Excel в папке
def process_all_files(directory, installation_name, control_date, option):
    # Получить список файлов Excel
    excel_files = get_excel_files(directory)

    # Инициализируем счетчик обработанных файлов
    processed_count = 0

    # Обрабатываем файлы
    for file_name in excel_files:
        file_path = os.path.join(directory, file_name)

        # В зависимости от выбранной опции (RGF или OEN) вызываем соответствующую функцию обработки
        if option == "RGF":
            process_excel_file_rgf(file_path, installation_name, control_date)
        elif option == "OEN":
            process_excel_file_oen(file_path, installation_name, control_date)
        else:
            continue  # Пропускаем файлы, если опция не указана

        # Увеличиваем счетчик обработанных файлов
        processed_count += 1

    # Логирование итогов
    logging.info(f'Обработано {processed_count} файлов.')
    
    # Создаем папку для сохранения обработанных файлов
    current_date = datetime.now().strftime("%d-%m-%Y")
    output_directory = f"Обработано_{current_date}"

    # Проверяем, существует ли уже папка с таким именем
    if os.path.exists(output_directory):
        # Если папка существует

        # Перемещаем обработанные файлы в существующую папку
        for file_name in excel_files:
            src_file_path = os.path.join(directory, file_name)
            dst_file_path = os.path.join(output_directory, file_name)
            shutil.move(src_file_path, dst_file_path)

        logging.info(f'ГОТОВЫЕ ФАЙЛЫ СОХРАНЕНЫ в существующей папке "{output_directory}".')

    else:
        # Создаем новую папку
        os.makedirs(output_directory)

        # Перемещаем обработанные файлы в новую папку
        for file_name in excel_files:
            src_file_path = os.path.join(directory, file_name)
            dst_file_path = os.path.join(output_directory, file_name)
            shutil.move(src_file_path, dst_file_path)

        logging.info(f'ГОТОВЫЕ ФАЙЛЫ СОХРАНЕНЫ в новой папке "{output_directory}".')

    return processed_count

# def process_excel_file(file_path, installation_name, control_date):
#     # Загрузите файл
#     workbook = load_workbook(file_path)

#     file_name = os.path.basename(file_path)

#     # Лист 'Характеристики'
#     characteristics_sheet = workbook['Характеристики']

#     # Установка значения в C2
#     characteristics_sheet['C2'].value = installation_name
#     logging.info(f'----- НАЧАЛО ОБРАБОТКИ ФАЙЛА: {file_name} -----"')
#     logging.info(f'Установлено значение {installation_name} в ячейку C2 на листе "Характеристики"')

#     # Проверка ячейки A1 и установка значения в C4
#     a1_value = characteristics_sheet['A1'].value
#     c4_value = None  # Значение по умолчанию для C4

#     if a1_value:
#         c4_value = a1_value
#         logging.info(f'Значение {a1_value} из ячейки A1 скопировано в ячейку C4 на листе "Характеристики"')
#     else:
#         logging.info(f'Ячейка A1 пуста, ячейка C4 будет установлена в None')

#     characteristics_sheet['C4'].value = c4_value

#     # Лист 'Диагностическая карта'
#     diagnostics_sheet = workbook['Диагностическая карта']

#     # Установка даты контроля в L5
#     diagnostics_sheet['L5'].value = control_date
#     logging.info(f'Установлено значение {control_date} в ячейку L5 на листе "Диагностическая карта"')

#     # Счетчик измененных строк в столбце A
#     changed_rows_count = 0

#     # Добавление формулы в столбец A
#     # Начинаем с первой строки данных (возможно, первая строка - это заголовок)
#     for row in range(8, diagnostics_sheet.max_row + 1):
#         # Формула для конкатенации B и C с точкой
#         formula = f'=CONCATENATE(B{row}, ".", C{row}, ".")'

#         # Проверка, что ячейки B{row} и C{row} не пусты перед установкой формулы в A{row}
#         if diagnostics_sheet[f'B{row}'].value is not None and diagnostics_sheet[f'C{row}'].value is not None:
#             diagnostics_sheet[f'A{row}'].value = formula
#             changed_rows_count += 1  # Увеличиваем счетчик для каждой измененной строки

#     # Логирование количества измененных строк
#     logging.info(f'Количество строк, где была установлена формула в столбец A: {changed_rows_count}')

#     # Сохранение файла
#     workbook.save(file_path)

#     # Логирование успешной обработки файла
#     logging.info(f'ФАЙЛ {file_name} ОБРАБОТАН УСПЕШНО!.')

#     # Загрузите файл
#     workbook = load_workbook(file_path)

#     # Лист 'Характеристики'
#     characteristics_sheet = workbook['Характеристики']

#     # Установка значения в C2
#     characteristics_sheet['C2'].value = installation_name

#     # Проверка ячейки A1 и установка значения в C4
#     if characteristics_sheet['A1'].value:
#         characteristics_sheet['C4'].value = characteristics_sheet['A1'].value
#     else:
#         characteristics_sheet['C4'].value = None

#     # Лист 'Диагностическая карта'
#     diagnostics_sheet = workbook['Диагностическая карта']

#     # Установка даты контроля в L5
#     diagnostics_sheet['L5'].value = control_date

#     # Сохранение файла
#     workbook.save(file_path)

#     # Логирование успешной обработки файла
#     # logging.info(f'Файл {file_path} обработан успешно.')

# def process_excel_file_oen(file_path, installation_name, control_date):
#     # Открываем файл Excel
#     workbook = load_workbook(file_path)

#     # Доступ к листу "УЗТ (коркарта)"
#     sheet = workbook['УЗТ (коркарта)']

#     # Проверка ячейки L5
#     cell_L5 = sheet['L5']
#     if cell_L5.value:
#         # Добавляем новый столбец перед B
#         sheet.insert_cols(2)

#         # Добавляем формулу в столбец B
#         for row in range(2, sheet.max_row + 1):
#             # Получаем ячейку в столбце 2
#             cell = sheet.cell(row=row, column=2)
#             # Проверяем, является ли ячейка объединенной
#             if isinstance(cell, MergedCell):
#                 # Найдем верхнюю левую ячейку объединенного диапазона
#                 top_left_cell = sheet.merged_cells.ranges.get(cell.coordinate)
#                 if top_left_cell:
#                     # Устанавливаем формулу в верхней левой ячейке
#                     top_left_cell.value = f'=CONCATENATE(B{row}; "."; C{row}; ".")'
#             else:
#                 # Устанавливаем значение, если ячейка не объединена
#                 cell.value = f'=CONCATENATE(B{row}; "."; C{row}; ".")'

#     # Вставляем введенное значение в ячейку F7
#     sheet['F7'].value = installation_name

#     # Проверка значения в ячейке F7 на соответствие формату ХХ.ХХ.ХХХХ
#     cell_value = sheet['F7'].value
#     pattern = re.compile(r'^\d{2}\.\d{2}\.\d{4}$')
#     if not pattern.match(cell_value):
#         # Если значение не соответствует формату, удаляем "г." из даты
#         sheet['F7'].value = cell_value.replace("г.", "")

#     # Сохраняем файл
#     workbook.save(file_path)

#     return f"Обработано успешно! Файл сохранен: {file_path}"

# def set_value_if_not_merged(sheet: Worksheet, row: int, col: int, value):
#     cell = sheet.cell(row=row, column=col)
#     merged_ranges = sheet.merged_cells.ranges

#     # Проверяем, находится ли ячейка в объединенном диапазоне
#     if any(range.min_row <= cell.row <= range.max_row and
#            range.min_col <= cell.column <= range.max_col
#            for range in merged_ranges):
#         # Находим верхний левый угол объединенного диапазона
#         for range in merged_ranges:
#             if range.min_row <= cell.row <= range.max_row and \
#                range.min_col <= cell.column <= range.max_col:
#                 # Устанавливаем значение в верхней левой ячейке
#                 top_left_cell = sheet.cell(row=range.min_row, column=range.min_col)
#                 top_left_cell.value = value
#                 break
#     else:
#         # Устанавливаем значение, если ячейка не объединена
#         cell.value = value




import os
import logging
import tkinter as tk
from tkinter import ttk, messagebox
from openpyxl import load_workbook
from datetime import datetime
import shutil


# # Настройка логгера
# logging.basicConfig(filename='processing.log', level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

log_filename = 'processing.log'

# Настройка логгера
def setup_logger():
    log_filename = 'processing.log'

    # Открытие файла в режиме 'a+' для добавления
    with open(log_filename, 'a+', encoding='utf-8') as file:
        file.seek(0)  # Перемещаемся в начало файла

        # Настройка логгера
        logging.basicConfig(
            level=logging.INFO,
            format='%(asctime)s - %(levelname)s - %(message)s',
            handlers=[
                logging.StreamHandler(),
                logging.FileHandler(log_filename, mode='a')  # Открытие лог-файла в режиме 'a' для добавления
            ]
        )
    # Открытие файла в режиме чтения и записи
    with open(log_filename, 'r+', encoding='utf-8') as file:
        # Чтение существующего содержимого файла
        existing_log_content = file.read()
        # Сброс позиции курсора в начало файла
        file.seek(0)

        # Настройка логгера для дописывания новой информации вверху файла
        logging.basicConfig(filename=log_filename, level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s', filemode='r+')

        # Пишем новые логи в лог-файл
        logging.info("New log entry")

        # Пишем существующее содержимое файла после новых логов
        file.write(existing_log_content)

# Вызов функции для настройки логгера
setup_logger()

# Функция для получения списка всех файлов Excel в папке
def get_excel_files(directory):
    excel_files = [file for file in os.listdir(directory) if file.lower().endswith(".xlsx")]
    logging.info(f'Найдены файлы Excel: {excel_files}')  # Логирование найденных файлов
    return excel_files


# Функция для обработки одного Excel-файла
def process_excel_file(file_path, installation_name, control_date):
    # Загрузите файл
    workbook = load_workbook(file_path)

    file_name = os.path.basename(file_path)

    # Лист 'Характеристики'
    characteristics_sheet = workbook['Характеристики']

    # Установка значения в C2
    characteristics_sheet['C2'].value = installation_name
    logging.info(f'----- НАЧАЛО ОБРАБОТКИ ФАЙЛА: {file_name} -----"')
    logging.info(f'Установлено значение {installation_name} в ячейку C2 на листе "Характеристики"')

    # Проверка ячейки A1 и установка значения в C4
    a1_value = characteristics_sheet['A1'].value
    c4_value = None  # Значение по умолчанию для C4

    if a1_value:
        c4_value = a1_value
        logging.info(f'Значение {a1_value} из ячейки A1 скопировано в ячейку C4 на листе "Характеристики"')
    else:
        logging.info(f'Ячейка A1 пуста, ячейка C4 будет установлена в None')

    characteristics_sheet['C4'].value = c4_value

    # Лист 'Диагностическая карта'
    diagnostics_sheet = workbook['Диагностическая карта']

    # Установка даты контроля в L5
    diagnostics_sheet['L5'].value = control_date
    logging.info(f'Установлено значение {control_date} в ячейку L5 на листе "Диагностическая карта"')

    # Счетчик измененных строк в столбце A
    changed_rows_count = 0

    # Добавление формулы в столбец A
    # Начинаем с первой строки данных (возможно, первая строка - это заголовок)
    for row in range(8, diagnostics_sheet.max_row + 1):
        # Формула для конкатенации B и C с точкой
        formula = f'=CONCATENATE(B{row}, ".", C{row}, ".")'

        # Проверка, что ячейки B{row} и C{row} не пусты перед установкой формулы в A{row}
        if diagnostics_sheet[f'B{row}'].value is not None and diagnostics_sheet[f'C{row}'].value is not None:
            diagnostics_sheet[f'A{row}'].value = formula
            changed_rows_count += 1  # Увеличиваем счетчик для каждой измененной строки

    # Логирование количества измененных строк
    logging.info(f'Количество строк, где была установлена формула в столбец A: {changed_rows_count}')

    # Сохранение файла
    workbook.save(file_path)

    # Логирование успешной обработки файла
    logging.info(f'ФАЙЛ {file_name} ОБРАБОТАН УСПЕШНО!.')

    # Загрузите файл
    workbook = load_workbook(file_path)

    # Лист 'Характеристики'
    characteristics_sheet = workbook['Характеристики']

    # Установка значения в C2
    characteristics_sheet['C2'].value = installation_name

    # Проверка ячейки A1 и установка значения в C4
    if characteristics_sheet['A1'].value:
        characteristics_sheet['C4'].value = characteristics_sheet['A1'].value
    else:
        characteristics_sheet['C4'].value = None

    # Лист 'Диагностическая карта'
    diagnostics_sheet = workbook['Диагностическая карта']

    # Установка даты контроля в L5
    diagnostics_sheet['L5'].value = control_date

    # Сохранение файла
    workbook.save(file_path)

    # Логирование успешной обработки файла
    # logging.info(f'Файл {file_path} обработан успешно.')

# Функция для обработки всех файлов Excel в папкеimport os
import shutil
from datetime import datetime

def process_all_files(directory, installation_name, control_date):
    # Получить список файлов Excel
    excel_files = get_excel_files(directory)

    # Инициализируем счетчик обработанных файлов
    processed_count = 0

    # Обрабатываем файлы
    for file_name in excel_files:
        file_path = os.path.join(directory, file_name)
        process_excel_file(file_path, installation_name, control_date)
        processed_count += 1

    # Логирование итогов
    logging.info(f'Обработано {processed_count} файлов.')

    # Создаем папку для сохранения обработанных файлов
    current_date = datetime.now().strftime('%d-%m-%Y')
    output_directory = f"Обработано_{current_date}"

    # Проверяем, существует ли уже папка с таким именем
    if os.path.exists(output_directory):
        # Если папка существует

        # Перемещаем обработанные файлы в существующую папку
        for file_name in excel_files:
            src_file_path = os.path.join(directory, file_name)
            dst_file_path = os.path.join(output_directory, file_name)
            shutil.move(src_file_path, dst_file_path)

        logging.info(f'ГОТОВЫЕ ФАЙЛЫ СОХРАНЕНЫ в существующей папке "{output_directory}".')

    else:
        # Создаем новую папку
        os.makedirs(output_directory)

        # Перемещаем обработанные файлы в новую папку
        for file_name in excel_files:
            src_file_path = os.path.join(directory, file_name)
            dst_file_path = os.path.join(output_directory, file_name)
            shutil.move(src_file_path, dst_file_path)

        logging.info(f'ГОТОВЫЕ ФАЙЛЫ СОХРАНЕНЫ в новой папке "{output_directory}".')

    return processed_count

# Функция для запуска обработки всех файлов по нажатию кнопки
def on_process_button_click():
    # Получаем введенные пользователем данные
    installation_name = installation_name_entry.get()
    control_date = control_date_entry.get()

    # Проверяем галочку РГФ
    if rgf_checkbox_var.get():
        # Проверка формата даты
        try:
            datetime.strptime(control_date, '%d.%m.%Y')
        except ValueError:
            messagebox.showerror("Ошибка", "Неправильный формат даты. Используйте формат DD.MM.YYYY.")
            return

        # Указываем путь к папке с файлами
        directory = os.getcwd()

        # Обрабатываем все файлы в папке и подсчитываем количество обработанных файлов
        file_count = process_all_files(directory, installation_name, control_date)

        # Сообщение о завершении обработки и вывод количества обработанных файлов
        messagebox.showinfo("УСПЕХ!", f"Обработка завершена успешно! \n Опция РГФ была выбрана. \n Обработано {file_count} файлов.")
         # Закрываем главное окно после отображения сообщения
        root.quit()
    else:
        # Обработка без опции РГФ
        # directory = os.getcwd()
        # file_count = process_all_files(directory, installation_name, control_date)

        # Сообщение о завершении обработки и вывод количества обработанных файлов
        messagebox.showinfo("Информация \n",f"Опция РГФ не выбрана. Обработано 0 файлов.")
         # Закрываем главное окно после отображения сообщения
        root.quit()
        logging.info('Опция РГФ не выбрана. Обработано 0 файлов.')

# Создаем главное окно приложения
root = tk.Tk()
root.title("GDC Excel Worker")

# Переменная для чекбокса РГФ
rgf_checkbox_var = tk.BooleanVar()

# Создаем переменную для чекбокса РГФ
rgf_checkbox_var = tk.IntVar(value=1)  # Установим значение 1, чтобы чекбокс был выбран по умолчанию

# Создаем чекбокс РГФ
rgf_checkbox = ttk.Checkbutton(root, text="РГФ", variable=rgf_checkbox_var)
rgf_checkbox.grid(row=0, column=0, padx=10, pady=10)

# Создаем метку и поле для ввода названия установки
ttk.Label(root, text="Введите название установки:").grid(row=1, column=0, padx=10, pady=10)
installation_name_entry = ttk.Entry(root)
installation_name_entry.grid(row=1, column=1, padx=10, pady=10)

# Создаем метку и поле для ввода даты контроля
ttk.Label(root, text="Введите дату контроля (ДД.ММ.ГГГГ):").grid(row=2, column=0, padx=10, pady=10)
control_date_entry = ttk.Entry(root)
control_date_entry.grid(row=2, column=1, padx=10, pady=10)

# Создаем кнопку для запуска обработки
process_button = ttk.Button(root, text="Запустить обработку", command=on_process_button_click)
process_button.grid(row=3, column=0, columnspan=2, pady=10)

# Запускаем главное окно приложения
root.mainloop()

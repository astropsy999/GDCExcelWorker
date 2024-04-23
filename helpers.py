from openpyxl.utils import get_column_letter

def insert_and_merge_column(sheet, col_index):
    """
    Вставляет столбец в определенном месте и объединяет ячейки в строках с 15 по 17.

    Параметры:
    sheet (Worksheet): Лист Excel, на котором производится вставка столбца.
    col_index (int): Индекс колонки, перед которой будет вставлен новый столбец (1 для столбца A, 2 для B и т.д.).
    """
    # Вставляем новый столбец перед столбцом col_index
    sheet.insert_cols(col_index)
    
    # Объединяем ячейки в диапазоне строк 15-17 для нового столбца
    start_row = 15
    end_row = 17
    start_col_letter = get_column_letter(col_index)
    end_col_letter = get_column_letter(col_index)
    
    # Объединяем ячейки в диапазоне строк 15-17 в новом столбце
    merged_range = f"{start_col_letter}{start_row}:{end_col_letter}{end_row}"
    sheet.merge_cells(merged_range)
    
    # # Устанавливаем значение для объединенных ячеек
    # sheet[f"{start_col_letter}{start_row}"].value = "Новый столбец"

def insert_column_in_range(sheet, col_index, start_row, end_row):
    """
    Вставляет столбец в определенном диапазоне строк на листе.

    Параметры:
    file_path (str): Путь к файлу Excel.
    sheet_name (str): Имя листа, на котором производится вставка столбца.
    col_index (int): Индекс колонки, перед которой будет вставлен новый столбец (1 для столбца A, 2 для B и т.д.).
    start_row (int): Начальная строка диапазона, в котором вставляется столбец.
    end_row (int): Конечная строка диапазона, в котором вставляется столбец.
    """

    header_shift = 4
  
     # Вставляем новый столбец в диапазоне строк
    for row in range(end_row, start_row -1 + header_shift, -1):
        # Сдвигаем значения в ячейках в строке вправо
        for col in range(sheet.max_column, col_index - 1, -1):
            # Копируем значение и стиль ячейки на одну колонку вправо
            source_cell = sheet.cell(row=row, column=col)
            target_cell = sheet.cell(row=row, column=col + 1)
            target_cell.value = source_cell.value
            target_cell._style = source_cell._style

        # Удаляем значение и стиль ячейки в вставляемом столбце
        source_cell = sheet.cell(row=row, column=col_index)
        source_cell.value = None
        source_cell._style = sheet.cell(row=row, column=col_index + 1)._style



# Найти начало таблицы
def find_table_start(sheet):
    """
    Функция ищет строку, содержащую текст "Результаты контроля:", и возвращает номер следующей строки.
    
    :param sheet: Лист Excel, который нужно проверить.
    :return: Номер строки, следующей за строкой с текстом "Результаты контроля:", или None, если строка не найдена.
    """
    # Перебираем строки листа
    for row_idx, row in enumerate(sheet.iter_rows(values_only=True), start=1):
        # Проверяем, содержит ли строка текст "Результаты контроля:"
        if row and any("Результаты контроля:" in str(cell) for cell in row):
            # Возвращаем номер следующей строки
            return row_idx + 1

    # Если строка с текстом "Результаты контроля:" не найдена
    print("Не удалось найти строку с 'Результаты контроля:'")
    return None

def find_table_end(sheet):
    """
    Функция ищет строку, содержащую текст "Примечание: Расположение зон контроля толщин стенок элементов трубопровода приведено на схеме контроля.",
    и возвращает номер предыдущей строки, которая считается концом таблицы.
    
    :param sheet: Лист Excel, который нужно проверить.
    :return: Номер строки, предшествующей строке с текстом "Примечание: Расположение зон контроля...", или None, если строка не найдена.
    """
    # Перебираем строки листа
    for row_idx, row in enumerate(sheet.iter_rows(values_only=True), start=1):
        # Проверяем, содержит ли строка текст "Примечание: Расположение зон контроля толщин стенок элементов трубопровода приведено на схеме контроля."
        if row and any("Примечание: Расположение зон контроля толщин стенок элементов трубопровода приведено на схеме контроля." in str(cell) for cell in row):
            # Возвращаем номер предыдущей строки, как конец таблицы
            return row_idx - 1

    # Если строка с текстом "Примечание: Расположение зон контроля..." не найдена
    print("Не удалось найти строку с 'Примечание: Расположение зон контроля...'")
    return None


# def is_date_time_format(cell_value):
#     """
#     Определяет, является ли значение ячейки строкой в формате даты 'YYYY-MM-DD HH:MM:SS'.

#     Параметры:
#         cell_value: Значение ячейки, которое необходимо проверить.

#     Возвращает:
#         True, если строка соответствует формату 'YYYY-MM-DD HH:MM:SS', иначе False.
#     """
#     date_time_format = '%Y-%m-%d %H:%M:%S'

#     try:
#         # Попытка разобрать строку как дату и время в указанном формате
#         datetime.strptime(cell_value, date_time_format)
#         return True
#     except ValueError:
#         # Если возникло исключение, значит строка не соответствует формату
#         return False

# def replace_formulas_with_values(sheet):
    """
    Функция заменяет формулы в ячейках их текущими значениями.
    """
    for row in sheet.iter_rows():

        for cell in row:
            cell_address = cell.coordinate

            formula = cell.value
            # value = cell.calculate(formula)
            if isinstance(formula, str) and formula.startswith('='):
                # Заменяем формулу текущим значением ячейки

                print('Формула', formula)
                print('Тип адреса', type(cell_address))
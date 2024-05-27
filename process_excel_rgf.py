import os
from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException
from utils import bcolors

def process_excel_file_rgf(file_path, installation_name, control_date):
    # Пытаемся загрузить файл, обрабатывая возможные ошибки
    try:
        workbook = load_workbook(file_path, read_only=False)
    except InvalidFileException:
        print(f'{bcolors.FAIL}Ошибка: файл {file_path} не является допустимым файлом Excel.{bcolors.ENDC}')
        return
    except Exception as e:
        print(f'{bcolors.FAIL}Ошибка при загрузке файла {file_path}: {str(e)}{bcolors.ENDC}')
        return

    file_name = os.path.basename(file_path)
    print(f'\n----- НАЧАЛО ОБРАБОТКИ ФАЙЛА: {file_name} -----')

    # Обрабатываем лист 'Характеристики'
    try:
        characteristics_sheet = workbook['Характеристики']
        # Устанавливаем значение в C2
        characteristics_sheet['C2'].value = installation_name # type: ignore
        print(f'Установлено значение {installation_name} в ячейку C2 на листе "Характеристики"')

        # Проверяем ячейку A1 и устанавливаем значение в C4
        a1_value = characteristics_sheet['A1'].value # type: ignore
        c4_value = a1_value if a1_value is not None else None

        characteristics_sheet['C4'].value = c4_value # type: ignore
        if a1_value:
            print(f'Значение {a1_value} из ячейки A1 скопировано в ячейку C4 на листе "Характеристики"')
        else:
            print('Ячейка A1 пуста, C4 установлено в None')

    except KeyError:
        
        raise KeyError(f'Лист "Характеристики" отсутствует в файле')
    finally:
        workbook.save(file_path)
        workbook.close()
        
    # Обрабатываем лист 'Диагностическая карта'
    try:
        diagnostics_sheet = workbook['Диагностическая карта']

        # Устанавливаем дату контроля в L5
        diagnostics_sheet['L5'].value = control_date # type: ignore
        print(f'Установлено значение {control_date} в ячейку L5 на листе "Диагностическая карта"')

        # Добавляем формулу в столбец A с проверкой пустых значений в B и C
        changed_rows_count = 0
        for row in range(8, diagnostics_sheet.max_row + 1): # type: ignore
            b_value = diagnostics_sheet[f'B{row}'].value # type: ignore
            c_value = diagnostics_sheet[f'C{row}'].value # type: ignore

            # Проверка на пустые значения B и C
            if b_value is not None and c_value is not None:
                formula = f'=CONCATENATE(B{row}, ".", C{row}, ".")'
                diagnostics_sheet[f'A{row}'].value = formula # type: ignore
                changed_rows_count += 1

        print(f'Количество строк, где была установлена формула в столбец A: {changed_rows_count}')

    except KeyError:

        raise KeyError(f'Лист "Диагностическая карта" отсутствует в файле')
    finally:
        workbook.save(file_path)
        workbook.close()
    # Сохраняем файл, обрабатывая возможные ошибки
    try:
        workbook.save(file_path)
        workbook.close()
        print(f'Файл {file_name} успешно обработан.')
    except PermissionError:
        print(f'{bcolors.FAIL}Ошибка: нет разрешения на запись в файл {file_name}.{bcolors.ENDC}')
    except Exception as e:
        print(f'{bcolors.FAIL}Ошибка при сохранении файла {file_name}: {str(e)}{bcolors.ENDC}')
    finally:
        workbook.save(file_path)
        workbook.close()